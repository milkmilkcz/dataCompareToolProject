"""
本地报告生成模块

提供本地报告生成功能：
- Markdown 格式报告
- CSV 格式报告
- Excel 格式报告（带完整格式化）
- JUnit XML 格式报告
- 结构化的输出
- 清晰优雅的格式
"""

import os
import csv
import logging
from datetime import datetime
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from ..core.validation import ValidationResult

logger = logging.getLogger(__name__)


class LocalReporter:
    """本地报告生成器"""

    def __init__(self):
        self.report_data = {
            'timestamp': datetime.now().isoformat(),
            'environment': 'UNKNOWN',
            'start_time': None,
            'end_time': None,
            'total_records': 0,
            'valid_records': 0,
            'error_records': 0,
            # 'skipped_records': 0,  # 删除跳过验证数
            'valid_details': [],
            'error_details': [],
            # 'skipped_details': [],  # 不再统计跳过验证的详情
            'status': 'UNKNOWN'
        }

    def set_environment(self, env: str):
        """设置环境"""
        self.report_data['environment'] = env.upper()

    def set_time_range(self, start_time: str, end_time: str):
        """设置时间范围"""
        self.report_data['start_time'] = start_time
        self.report_data['end_time'] = end_time

    def add_validation_result(self, result: ValidationResult):
        """添加验证结果"""
        summary = result.get_summary()
        self.report_data['total_records'] = summary['total']
        self.report_data['valid_records'] = summary['valid_count']
        self.report_data['error_records'] = summary['error_count']
        # self.report_data['skipped_records'] = summary['skipped_count']  # 删除跳过验证数

        # 添加有效记录详情
        for record in result.valid_records:
            self.report_data['valid_details'].append({
                'policyNum': str(record.get('policyNum', 'N/A')),
                'msgHead': str(record.get('msgHead', 'N/A')),
                'bizTime': str(record.get('bizTime', 'N/A')),
                'status': str(record.get('status', 'N/A'))
            })

        # 添加错误详情
        for error_item in result.error_records:
            record = error_item['record']
            self.report_data['error_details'].append({
                'policyNum': str(record.get('policyNum', 'N/A')),
                'msgHead': str(record.get('msgHead', 'N/A')),
                'bizTime': str(record.get('bizTime', 'N/A')),
                'status': str(record.get('status', 'N/A')),
                'error_reason': error_item['error_reason']
            })

        # 不再统计和输出 skipped_details
        # for skipped_item in result.skipped_records:
        #     ...

        # 确定状态 - 跳过验证的记录不影响状态，只有真正的错误才算失败
        if self.report_data['error_records'] > 0:
            self.report_data['status'] = 'FAILURE'
        else:
            self.report_data['status'] = 'SUCCESS'

    def generate_markdown_report(self) -> str:
        """生成 Markdown 格式的报告"""
        data = self.report_data

        # 标题
        report = "# 数据验证报告\n\n"

        # 基本信息
        report += "## 基本信息\n\n"
        report += f"- **生成时间**: {data['timestamp']}\n"
        report += f"- **环境**: {data['environment']}\n"
        report += f"- **时间范围**: {data['start_time']} ~ {data['end_time']}\n"
        report += f"- **状态**: {'✅ 成功' if data['status'] == 'SUCCESS' else '❌ 失败'}\n\n"

        # 统计信息
        report += "## 统计信息\n\n"
        report += f"- **总记录数**: {data['total_records']}\n"
        report += f"- **有效记录数**: {data['valid_records']}\n"
        report += f"- **错误记录数**: {data['error_records']}\n\n"
        # 不再输出跳过验证数
        # report += f"- **跳过验证数**: {data['skipped_records']}\n\n"

        # 有效记录详情
        if data['valid_details']:
            report += "## 有效记录详情\n\n"
            report += "| 序号 | 保单号 | 业务表 | 业务时间 | 状态 |\n"
            report += "|------|--------|--------|----------|------|\n"

            for i, valid in enumerate(data['valid_details'], 1):
                report += f"| {i} | {valid['policyNum']} | {valid['msgHead']} | {valid['bizTime']} | {valid['status']} |\n"

            report += "\n"

        # 错误详情
        if data['error_details']:
            report += "## 错误详情\n\n"
            report += "| 序号 | 保单号 | 业务表 | 业务时间 | 错误原因 |\n"
            report += "|------|--------|--------|----------|----------|\n"

            for i, error in enumerate(data['error_details'], 1):
                report += f"| {i} | {error['policyNum']} | {error['msgHead']} | {error['bizTime']} | {error['error_reason']} |\n"

            report += "\n"

        # 总结
        report += "## 总结\n\n"
        if data['status'] == 'SUCCESS':
            report += "✅ 数据验证通过，所有记录均有效。\n"
        else:
            report += f"❌ 数据验证失败，发现 {data['error_records']} 条错误记录。\n"

        return report

    def export_markdown(self, filename: str) -> bool:
        """导出为 Markdown 格式"""
        try:
            report_content = self.generate_markdown_report()
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(report_content)
            logger.info(f"本地报告已导出到 {filename}")
            return True
        except Exception as e:
            logger.error(f"导出 Markdown 报告失败: {e}")
            return False

    def export_junit_xml(self, filename: str) -> bool:
        """
        导出为 JUnit XML 格式（用于CI/CD集成）

        JUnit XML 格式便于 Jenkins 等CI工具解析测试结果
        """
        try:
            data = self.report_data

            # 创建根元素
            testsuite = ET.Element('testsuite')
            testsuite.set('name', f'DataValidation_{data["environment"]}')
            testsuite.set('tests', str(data['total_records']))
            testsuite.set('failures', str(data['error_records']))
            testsuite.set('timestamp', data['timestamp'])

            # 添加有效记录的测试用例
            if data['valid_records'] > 0:
                testcase = ET.SubElement(testsuite, 'testcase')
                testcase.set('name', f"Valid Records ({data['valid_records']})")
                testcase.set('classname', 'DataValidation.ValidRecords')
                testcase.set('time', '0.001')  # 模拟执行时间

            # 添加错误记录的测试用例
            for i, error in enumerate(data['error_details'], 1):
                testcase = ET.SubElement(testsuite, 'testcase')
                testcase.set('name', f"Error Record {i}: {error['policyNum']}")
                testcase.set('classname', 'DataValidation.ErrorRecords')
                testcase.set('time', '0.001')

                # 添加失败信息
                failure = ET.SubElement(testcase, 'failure')
                failure.set('type', 'ValidationError')
                failure.text = (
                    f"Policy: {error['policyNum']}\n"
                    f"Table: {error['msgHead']}\n"
                    f"Time: {error['bizTime']}\n"
                    f"Reason: {error['error_reason']}"
                )

            # 保存 XML
            tree = ET.ElementTree(testsuite)
            tree.write(filename, encoding='utf-8', xml_declaration=True)
            logger.info(f"JUnit XML 报告已导出到 {filename}")
            return True
        except Exception as e:
            logger.error(f"导出 JUnit XML 失败: {e}")
            return False

    def export_excel(self, filename: str) -> bool:
        """
        导出为 Excel 格式（带完整格式化）

        包含：
        - 边框样式
        - 标题加粗增大字号
        - 列对齐设置
        - 自适应列宽
        - 错误原因自动换行
        """
        try:
            data = self.report_data

            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "数据验证报告"

            # 定义样式
            # 边框样式
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 标题样式（加粗，增大字号，居中）
            title_font = Font(bold=True, size=14)
            title_alignment = Alignment(horizontal='center', vertical='center')

            # 表头样式（加粗，居中）
            header_font = Font(bold=True, size=12)
            header_alignment = Alignment(horizontal='center', vertical='center')
            header_fill = PatternFill(start_color='FFE6E6FA', end_color='FFE6E6FA', fill_type='solid')

            # 数据样式（居中对齐）
            data_alignment = Alignment(horizontal='center', vertical='center')

            # 错误原因样式（左对齐，自动换行）
            error_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            # 当前行号
            current_row = 1

            # 写入标题
            ws.merge_cells(f'A{current_row}:E{current_row}')
            title_cell = ws[f'A{current_row}']
            title_cell.value = "数据验证报告"
            title_cell.font = title_font
            title_cell.alignment = title_alignment
            title_cell.border = thin_border
            current_row += 1

            # 空行
            current_row += 1

            # 写入基本信息标题
            ws.merge_cells(f'A{current_row}:E{current_row}')
            info_title_cell = ws[f'A{current_row}']
            info_title_cell.value = "基本信息"
            info_title_cell.font = Font(bold=True, size=12)
            info_title_cell.alignment = Alignment(horizontal='left')
            current_row += 1

            # 基本信息数据
            info_data = [
                ["生成时间", data['timestamp']],
                ["环境", data['environment']],
                ["时间范围", f"{data['start_time']} ~ {data['end_time']}"],
                ["状态", "✅ 成功" if data['status'] == 'SUCCESS' else "❌ 失败"]
            ]

            for label, value in info_data:
                ws[f'A{current_row}'] = label
                ws[f'A{current_row}'].font = Font(bold=True)
                ws[f'A{current_row}'].border = thin_border
                ws[f'A{current_row}'].alignment = data_alignment

                ws[f'B{current_row}'] = value
                ws[f'B{current_row}'].border = thin_border
                ws[f'B{current_row}'].alignment = data_alignment
                current_row += 1

            # 空行
            current_row += 1

            # 写入统计信息标题
            ws.merge_cells(f'A{current_row}:E{current_row}')
            stat_title_cell = ws[f'A{current_row}']
            stat_title_cell.value = "统计信息"
            stat_title_cell.font = Font(bold=True, size=12)
            stat_title_cell.alignment = Alignment(horizontal='left')
            current_row += 1

            # 统计信息数据
            stat_data = [
                ["总记录数", data['total_records']],
                ["有效记录数", data['valid_records']],
                ["错误记录数", data['error_records']],
                # ["跳过验证数", data['skipped_records']]  # 不再输出跳过验证数
            ]

            for label, value in stat_data:
                ws[f'A{current_row}'] = label
                ws[f'A{current_row}'].font = Font(bold=True)
                ws[f'A{current_row}'].border = thin_border
                ws[f'A{current_row}'].alignment = data_alignment

                ws[f'B{current_row}'] = value
                ws[f'B{current_row}'].border = thin_border
                ws[f'B{current_row}'].alignment = data_alignment
                current_row += 1

            # 如果有错误详情，写入错误详情表
            if data['error_details']:
                # 空行
                current_row += 1

                # 错误详情标题
                ws.merge_cells(f'A{current_row}:E{current_row}')
                error_title_cell = ws[f'A{current_row}']
                error_title_cell.value = "错误详情"
                error_title_cell.font = Font(bold=True, size=12)
                error_title_cell.alignment = Alignment(horizontal='left')
                current_row += 1

                # 表头
                headers = ["序号", "保单号", "业务表", "业务时间", "错误原因"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.value = header
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = thin_border
                    cell.fill = header_fill
                current_row += 1

                # 错误数据
                for i, error in enumerate(data['error_details'], 1):
                    # 序号
                    cell = ws.cell(row=current_row, column=1)
                    cell.value = i
                    cell.alignment = data_alignment
                    cell.border = thin_border

                    # 保单号
                    cell = ws.cell(row=current_row, column=2)
                    cell.value = error['policyNum']
                    cell.alignment = data_alignment
                    cell.border = thin_border

                    # 业务表
                    cell = ws.cell(row=current_row, column=3)
                    cell.value = error['msgHead']
                    cell.alignment = data_alignment
                    cell.border = thin_border

                    # 业务时间
                    cell = ws.cell(row=current_row, column=4)
                    cell.value = error['bizTime']
                    cell.alignment = data_alignment
                    cell.border = thin_border

                    # 错误原因 - 特殊处理，自动换行，每条错误单独一行
                    cell = ws.cell(row=current_row, column=5)
                    error_reason = str(error['error_reason'])  # 确保是字符串

                    # 处理复杂的错误信息，按分号分割成多行
                    if ';' in error_reason:
                        error_lines = [line.strip() for line in error_reason.split(';') if line.strip()]
                        cell.value = '\n'.join(error_lines)
                    else:
                        cell.value = error_reason

                    cell.alignment = error_alignment
                    cell.border = thin_border

                    current_row += 1

            # 设置列宽自适应
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        cell_value = cell.value
                        if cell_value is not None:
                            # 处理多行文本，计算最长行的长度
                            value_str = str(cell_value)
                            if '\n' in value_str:
                                lines = value_str.split('\n')
                                line_lengths = [len(line) for line in lines]
                                max_line_length = max(line_lengths) if line_lengths else 0
                                max_length = max(max_length, max_line_length)
                            else:
                                max_length = max(max_length, len(value_str))
                    except Exception:
                        pass

                # 设置最小宽度和最大宽度
                if column[0].column == 5:  # 错误原因列
                    adjusted_width = min(max_length * 0.7, 80)  # 错误原因列可以更宽
                else:
                    adjusted_width = min(max_length * 1.2, 25)  # 其他列适中宽度

                ws.column_dimensions[column_letter].width = max(adjusted_width, 10)

            # 保存文件
            wb.save(filename)
            logger.info(f"Excel 报告已导出到 {filename}")
            return True

        except Exception as e:
            logger.error(f"导出 Excel 报告失败: {e}")
            return False


def create_local_report(result: ValidationResult, env: str, start_time: str, end_time: str, reports_dir: str = 'reports') -> bool:
    """
    快速函数：创建本地多种格式的报告（Markdown、CSV、Excel、JUnit XML）

    Args:
        result: ValidationResult 对象
        env: 环境名称
        start_time: 开始时间
        end_time: 结束时间
        reports_dir: 报告目录

    Returns:
        是否成功
    """
    # 确保报告目录存在
    os.makedirs(reports_dir, exist_ok=True)

    # 生成时间戳
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    # 创建报告器
    reporter = LocalReporter()
    reporter.set_environment(env)
    reporter.set_time_range(start_time, end_time)
    reporter.add_validation_result(result)

    success = True

    # 生成 Markdown 报告
    md_filename = os.path.join(reports_dir, f'data_validation_report_{timestamp}.md')
    if not reporter.export_markdown(md_filename):
        success = False

    # 生成 CSV 报告
    csv_filename = os.path.join(reports_dir, f'data_validation_report_{timestamp}.csv')
    try:
        # 直接生成CSV内容并写入
        data = reporter.report_data
        csv_lines = []

        # 写入头部
        csv_lines.append(["生成时间", "环境", "时间范围", "状态", "总记录数", "有效记录数", "错误记录数"])
        csv_lines.append([
            data['timestamp'],
            data['environment'],
            f"{data['start_time']} ~ {data['end_time']}",
            "成功" if data['status'] == 'SUCCESS' else "失败",
            data['total_records'],
            data['valid_records'],
            data['error_records'],
        ])

        # 错误详情
        if data['error_details']:
            csv_lines.append([])
            csv_lines.append(["序号", "保单号", "业务表", "业务时间", "错误原因"])
            for i, error in enumerate(data['error_details'], 1):
                csv_lines.append([i, error['policyNum'], error['msgHead'], error['bizTime'], error['error_reason']])

        # 写入文件 - 使用UTF-8 with BOM，确保Excel正确显示中文
        with open(csv_filename, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerows(csv_lines)
        logger.info(f"CSV 报告已导出到 {csv_filename} (UTF-8 with BOM)")
    except Exception as e:
        logger.error(f"导出 CSV 报告失败: {e}")
        success = False

    # 生成 JUnit XML 报告
    junit_filename = os.path.join(reports_dir, f'data_validation_report_{timestamp}.xml')
    if not reporter.export_junit_xml(junit_filename):
        success = False

    # 生成 Excel 报告
    excel_filename = os.path.join(reports_dir, f'data_validation_report_{timestamp}.xlsx')
    if not reporter.export_excel(excel_filename):
        success = False

    if success:
        logger.info(f"所有报告已生成并保存到: {reports_dir}")
        print(f"报告已生成：")
        print(f"  Markdown: {md_filename}")
        print(f"  CSV: {csv_filename}")
        print(f"  JUnit XML: {junit_filename}")
        print(f"  Excel: {excel_filename}")

    return success
